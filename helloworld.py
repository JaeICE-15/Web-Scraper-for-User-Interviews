from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import time

#preventing the driver from closing every time it completes an instruction
options = Options()
options.add_experimental_option("detach", True)

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get("https://www.userinterviews.com/studies?sort=-id")

#automatically maximises the window when run
driver.maximize_window()

#clicks the filters on the page to display only online and over-the-phone listings
filt1 = driver.find_element("xpath", "//input[@id='interview-type-2']").click()
time.sleep(2)
filt2 = driver.find_element("xpath", "//input[@id='interview-type-4']").click()

#waits for maximum of 200 seconds for the whole page to finish loading
driver.implicitly_wait(200)
#scrolls to the bottom of the page
driver.execute_script("window.scrollTo(0,document.body.scrollHeight,)")

#searches for links that lead to project listings
time.sleep(16)
links = driver.find_elements("xpath","//a[text()[contains(., 'Apply')]]")

project_list = []
description_list = []
reward_list = []

#compiles all project listing links into a list
for link in links:
    project_list.append(link.get_attribute("href"))

#assigns the current tab as the original
original_tab = driver.current_window_handle

#visits each listing in the project_list to scrape each listing's description and reward into their respective lists
for project in range(len(project_list)):
    driver.switch_to.new_window('tab')
    driver.get(project_list[project])
    description = driver.find_element("xpath", "//*[@id='ui-window-root']/div/section/dl/div[5]/div/div[1]/dd")
    description_list.append(BeautifulSoup(description.get_attribute("innerHTML"), "lxml").text)
    reward = driver.find_element("xpath", "//*[@id='ui-window-root']/div/section/dl/div[3]/dd")
    reward_list.append(BeautifulSoup(reward.get_attribute("innerHTML"), "lxml").text)
    driver.close()
    driver.switch_to.window(original_tab)

#loads the excel spreadsheet file
file_path = "my_list.xlsx"
wb = load_workbook(file_path)
ws = wb.active

#writes the contents of the lists into 3 different columns in the excel spreadsheet
for index, value in enumerate(project_list, start=1):
    ws.cell(row=index, column=1, value=value)

for index, value in enumerate(description_list, start=1):
    ws.cell(row=index, column=2, value=value)

for index, value in enumerate(reward_list, start=1):
    ws.cell(row=index, column=3, value=value)

#saves the changes
wb.save(file_path)
print("Data written and saved successfully")